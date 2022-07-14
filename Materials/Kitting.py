import pyodbc
import numpy as np
import os
import win32print
import win32api
import openpyxl
from openpyxl import load_workbook
from tkinter import simpledialog
from tkinter import messagebox
from tkinter import NORMAL
from tkinter import END
from tkinter import DISABLED

def KittingLabel(Report_Text,Book):
	conn2 = pyodbc.connect("Driver={ODBC Driver 17 for SQL Server};"
	                           "Server=DC01;"
	                           "Database=Interactive;"
	                           "Trusted_Connection=yes;")
	c = conn2.cursor()

	WO = simpledialog.askstring("WONO", "Enter the Work Order Number")

	#if user clicks cancel or X, end function
	if WO == None:
		pass

	else:
		try:
			#fetch all the data for that WO
			EXE_STRING = "EXEC Interactive.dbo.MAT_KittingLabel @iwo=?"
			c.execute(EXE_STRING,WO)
			data = c.fetchall()
			#if no data found for that work order, give an error and ask again
			if not data:
				messagebox.showerror("Kitting Labels", "Work order not found.")
				KittingLabel(Report_Text,Book)
			else:
				np_data = np.array(data)

				if np_data[0,3] == 'LEADSPCL' or np_data[0,3] == 'LF SPCL':
					#open the template excel file
					xfile = openpyxl.load_workbook('H:\\_LABELS\\KITTING_BIN_LABELS\\FM-649 Production Kit Label-SPCL-REV_4.xlsx')
				else:
					#open the template excel file
					xfile = openpyxl.load_workbook('H:\\_LABELS\\KITTING_BIN_LABELS\\FM-649 Production Kit Label-REV_4.xlsx')

				#since this information is always the same no matter how many work centers, just insert the first value of the array
				sheet = xfile["Sheet1"]
				sheet['A2'] = 'WO #: ' + np_data[0,0]
				sheet['A4'] = 'Assembly #: ' + np_data[0,1]
				sheet['C4'] = 'Rev: ' + np_data[0,2]
				sheet['A6'] = 'Customer Name: ' + np_data[0,4]
				#limit this to 40 characters to prevent printing a 2nd page
				sheet['A7'] = 'Job Name: ' + np_data[0,5][:41]
				#must convert data type DATE to string
				sheet['A9'] = 'Materials Due Date: ' + str(np_data[0,6])
				sheet['A11'] = 'Work Center: '
				sheet['A3'] = str(np_data[0,8])
				#x = row number
				x = 12
				y = 12
				rowcount = 0

				for row in np_data:
					if rowcount >= 5:
						#insert insert into cell A[y], places each work center on a new row
						cell = 'B' + str(y)
						#place a checkbox after each work center
						sheet[cell] = np_data[rowcount,7] + "[ ]"
						rowcount += 1
						y += 1
					else:
						#insert insert into cell A[x], places each work center on a new row
						cell = 'A' + str(x)
						#place a checkbox after each work center
						sheet[cell] = np_data[rowcount,7] + "[ ]"
						rowcount += 1
						x += 1


				sheet['A18'] = ('Bin:    /')
				#save the file to the finished sheet
				xfile.save('C:\\Users\\' + os.getlogin() + '\\Desktop\\KittingLabel.xlsx')
				#if assembly requires special instruction, print to a separate printer
				if np_data[0,3] == 'LEADSPCL' or np_data[0,3] == 'LF SPCL':
					Book.update()
					#ask how many copies they would like to print
					pages = simpledialog.askstring("Copies", "How Many Copies would you like?")
					#if entry box is empty, or cancel or X is clicked, do not print and say no page count selected
					if pages == None or pages == '':
						Report_Text.configure(state = NORMAL)
						Report_Text.delete('1.0',END)
						Report_Text.insert(END, "No Page Count Selected\n")
						Report_Text.configure(state = DISABLED)
						Report_Text.see("end")
					elif int(pages) > 10:
						messagebox.showerror("Kitting Labels", "Cannot print more than 10 pages.")
					else:
						#print the finished sheet a number of times equal to the input number of copies
						for x in range(0,int(pages)):
							win32api.ShellExecute (
							  0,
							  "printto",
							  #print the finished copy
							  'C:\\Users\\' + os.getlogin() + '\\Desktop\\KittingLabel.xlsx',
							  #
							  # If this is None, the default printer will
							  # be used anyway.
							  #
							  #specify the name of the printer
							  '"%s"' % "TASKalfa 3051ci",
							  ".",
							  0
							)
						Report_Text.configure(state = NORMAL)
						Report_Text.delete('1.0',END)
						#tell the user where it was printed to
						Report_Text.insert(END, pages + " Pages Printed to Kyocera 3051ci-HUB Tray 2\nSPECIAL HANDLING")
						Report_Text.configure(state = DISABLED)
						Report_Text.see("end")

				else:
					Book.update()
					pages = simpledialog.askstring("Copies", "How Many Copies would you like?")
					if pages == None or pages == '':
						Report_Text.configure(state = NORMAL)
						Report_Text.delete('1.0',END)
						Report_Text.insert(END, "No Page Count Selected\n")
						Report_Text.configure(state = DISABLED)
						Report_Text.see("end")
					elif int(pages) > 10:
						messagebox.showerror("Kitting Labels", "Cannot print more than 10 pages.")
					else:
						for x in range(0,int(pages)):

							win32api.ShellExecute (
							  0,
							  "printto",
							  'C:\\Users\\' + os.getlogin() + '\\Desktop\\KittingLabel.xlsx',
							  #
							  # If this is None, the default printer will
							  # be used anyway.
							  #
							  '"%s"' % "\\\\Server2012\\Kyocera 3051ci-HUB",
							  ".",
							  0
							)

						Report_Text.configure(state = NORMAL)
						Report_Text.delete('1.0',END)
						Report_Text.insert(END, pages + " Pages Printed to Kyocera 3051ci-HUB Tray 1\n")
						Report_Text.configure(state = DISABLED)
						Report_Text.see("end")

		except Exception as e:
			e1 = str(e)
			Report_Text.configure(state = NORMAL)
			Report_Text.delete('1.0',END)
			Report_Text.insert(END, e1 + "\n")
			Report_Text.configure(state = DISABLED)
			Report_Text.see("end")



				